# -*- coding: utf-8 -*-
#
# author: amanul haque
#

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import re
import html
import pandas as pd
from bs4 import BeautifulSoup as bs
from pprint import pprint
import random
import chardet
import numpy as np
from unidecode import unidecode

class data_preprocessing_1:
    
    def __init__(self):
        
        self.input_data_file_path = 'final_data/Heckman2015_Piazza.xlsx'
        self.output_file = 'final_data/processed_output.csv'
    
    def read_file(self, filepath):
        
        wb=load_workbook(filepath)
        # select demo.xlsx
        sheet=wb.active
        # get max row count
        max_row=sheet.max_row
        # get max column count
        max_column=sheet.max_column
        # iterate over all cells 
        # iterate over all rows
        for i in range(1,max_row+1):
             
             # iterate over all columns
             for j in range(1,max_column+1):
                  # get particular cell value    
                  cell_obj=sheet.cell(row=i,column=j)
                  # print cell value     
                  #print(cell_obj.value,end=' | ')
             # print new line
             #print('\n')
                
    
    def read_data(self, filepath):
        
        data = pd.read_excel(filepath, sep=',')
        return data
    
    def remove_encoding_from_text(self, cmt):
        
        return re.sub(r'[^\x00-\x7F]+',' ', cmt)
        '''
        return ''.join([i if ord(i) < 128 else ' ' for i in cmt])
        newString = (cmt.encode('ascii', errors='ignore')).decode("utf-8")
        print("Even before ", unidecode(cmt))
        return newString
        '''
        
    def process_html_tags(self, cmt):
        
        if(cmt != ''):
            soup = bs(cmt, 'lxml')
            cmt_text = soup.text
            if cmt_text != None:
                return cmt_text
            else:
                return 'na'
        else:
            return 'na'

    def cast_timestamp(self, data):
        
        time_list = data['Time']
        timestamps_list = []
        for time in time_list:
            timestamp = pd.Timestamp(time)
            timestamps_list.append(timestamp)
        
        data = data.drop(columns = 'Time', axis = 1)
        data['Time'] = timestamps_list
        return data
    
    def process_data(self, X):
        
        new_X = [] 
        for cmt in X:
            #print(row['Content'])
            cmt_processed = self.process_html_tags(str(cmt))
            #print("Before : ", cmt_processed)
            #print()
            cmt_processed = self.remove_encoding_from_text(cmt_processed)
            #print("processed_comments : ", cmt_processed)
            new_X.append(cmt_processed)
        
        return np.array(new_X)
        
    
    
    def process_data_2(self, df):
        
        i = 0
        column_headers = df.columns.values
        df_new = pd.DataFrame(columns=column_headers)
        for index, row in df.iterrows():
            #print(row['Content'])
            cmt_processed = self.process_html_tags(str(row['Content']))
            #print("Before : ", cmt_processed)
            #print()
            cmt_processed = self.remove_encoding_from_text(cmt_processed)
            #print("processed_comments : ", cmt_processed)
            row['Content'] = cmt_processed
            df_new.loc[i] = row
            i += 1
        return df_new
        
    
if __name__ == '__main__':
    
    dp = data_preprocessing_1()
    df = dp.read_data(dp.input_data_file_path)
    df = dp.process_data_2(df)
    #print(df)
    df.to_csv(dp.output_file, sep=',')
    #print(df)
    
    
    
    